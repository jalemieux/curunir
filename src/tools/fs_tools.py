import glob as glob_module
import re
import subprocess
from pathlib import Path

from src.config import AgentConfig


class JailError(Exception):
    """Raised when a path escapes the persona's FS sandbox (``config.workdir``)."""


def _jail_root(config: AgentConfig) -> Path | None:
    """Return the realpath'd sandbox root, or None when jailing is disabled.

    Isolation between personas is the #420 hard requirement. In production each
    persona's FS-touching tools run inside a container/namespace whose mount
    root is this directory; here (and as defense-in-depth inside that
    container) we enforce a hardened realpath path-jail so persona A's tools
    can never reach persona B's context/<persona>/ tree.
    """
    if not getattr(config, "fs_jail", False):
        return None
    return Path(config.workdir).resolve()


def _resolve_in_jail(config: AgentConfig, raw: str) -> Path:
    """Resolve *raw* (relative to the jail root) and assert containment.

    Relative paths anchor at ``config.workdir`` (not the process cwd) so the
    model addresses its own sandbox naturally. The final realpath — symlinks
    resolved — must live under the jail root; ``..``, absolute escapes, and
    symlink escapes all fail closed via ``Path.relative_to``. When jailing is
    disabled the raw path is returned untouched (legacy single-tenant
    behavior).
    """
    root = _jail_root(config)
    if root is None:
        return Path(raw)
    p = Path(raw)
    if not p.is_absolute():
        p = root / p
    resolved = p.resolve()
    if resolved != root and root not in resolved.parents:
        raise JailError(
            f"path {raw!r} escapes the persona sandbox; FS tools are confined "
            f"to {root}"
        )
    return resolved


def exec_glob(args: dict, config: AgentConfig) -> str:
    """Find files matching a glob pattern."""
    try:
        pattern = args["pattern"]
        root = args.get("path", ".")
        # Strip leading slashes so patterns are always relative to root_dir.
        # An absolute pattern like "/**/*.py" would bypass root_dir and scan
        # the entire filesystem, hanging the process.
        pattern = pattern.lstrip("/")
        if not pattern:
            return "Error: empty glob pattern"
        root = str(_resolve_in_jail(config, root))
        matches = glob_module.glob(pattern, root_dir=root, recursive=True)
        matches = sorted(matches)[:1000]
        return "\n".join(matches)
    except JailError as e:
        return f"Error: {e}"
    except Exception as e:
        return f"Error: {e}"


def _find_rg() -> str | None:
    """Find the ripgrep binary path."""
    import shutil
    return shutil.which("rg")


def exec_grep(args: dict, config: AgentConfig) -> str:
    """Search file contents using ripgrep if available, else pure Python."""
    pattern = args["pattern"]
    path = args.get("path", ".")
    glob_filter = args.get("glob")
    output_mode = args.get("output_mode", "content")
    context_lines = args.get("context", 0)

    try:
        path = str(_resolve_in_jail(config, path))
    except JailError as e:
        return f"Error: {e}"

    rg_path = _find_rg()
    if rg_path:
        try:
            cmd = [rg_path, "--no-heading"]
            if glob_filter:
                cmd.extend(["--glob", glob_filter])
            if output_mode == "files_with_matches":
                cmd.append("--files-with-matches")
            elif output_mode == "count":
                cmd.append("--count")
            if context_lines:
                cmd.extend(["-C", str(context_lines)])
            cmd.extend([pattern, path])
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            return result.stdout if result.stdout else ""
        except Exception as e:
            return f"Error: {e}"

    # Pure Python fallback
    try:
        regex = re.compile(pattern)
        root = Path(path)

        # Collect files to search
        if root.is_file():
            files = [root]
        else:
            if glob_filter:
                files = list(root.glob(glob_filter))
            else:
                files = [f for f in root.rglob("*") if f.is_file()]

        output_lines = []
        for filepath in sorted(files):
            try:
                text = filepath.read_text(errors="replace")
            except Exception:
                continue

            lines = text.splitlines()
            matching_indices = [i for i, line in enumerate(lines) if regex.search(line)]

            if not matching_indices:
                continue

            if output_mode == "files_with_matches":
                output_lines.append(str(filepath))
            elif output_mode == "count":
                output_lines.append(f"{filepath}:{len(matching_indices)}")
            else:
                # content mode with optional context
                shown: set[int] = set()
                for idx in matching_indices:
                    for j in range(max(0, idx - context_lines), min(len(lines), idx + context_lines + 1)):
                        shown.add(j)
                for j in sorted(shown):
                    output_lines.append(f"{filepath}:{j + 1}:{lines[j]}")

        return "\n".join(output_lines)
    except Exception as e:
        return f"Error: {e}"


def _read_pdf(path: Path) -> str:
    import pymupdf
    doc = pymupdf.open(str(path))
    pages = []
    for i, page in enumerate(doc, 1):
        text = page.get_text()
        if text.strip():
            pages.append(f"--- Page {i} ---\n{text}")
    doc.close()
    return "\n".join(pages) if pages else "(no text content found in PDF)"


def _read_docx(path: Path) -> str:
    import docx
    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def _read_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append("\t".join(str(c) if c is not None else "" for c in row))
        if rows:
            parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n".join(parts) if parts else "(empty spreadsheet)"


def _read_csv(path: Path) -> str:
    import csv
    with open(path, newline="", errors="replace") as f:
        reader = csv.reader(f)
        rows = ["\t".join(row) for row in reader]
    return "\n".join(rows)


_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".svg", ".ico"}

_BINARY_READERS: dict[str, callable] = {
    ".pdf": _read_pdf,
    ".docx": _read_docx,
    ".xlsx": _read_xlsx,
    ".xls": _read_xlsx,
    ".csv": _read_csv,
}


def exec_read(args: dict, config: AgentConfig) -> str:
    """Read a file. Handles text, PDF, DOCX, XLSX, and CSV."""
    try:
        path = _resolve_in_jail(config, args["file_path"])
        if not path.exists():
            return f"Error: File not found: {path}"

        suffix = path.suffix.lower()
        reader = _BINARY_READERS.get(suffix)
        if reader:
            return reader(path)

        if suffix in _IMAGE_EXTENSIONS:
            return f"This is an image file ({suffix}). Image content was already provided inline when the email was received. Use the information from the original message to respond."

        lines = path.read_text().splitlines()
        offset = args.get("offset", 1)
        limit = args.get("limit", len(lines))

        # offset is 1-based
        start = max(0, offset - 1)
        end = start + limit
        selected = lines[start:end]

        numbered = []
        for i, line in enumerate(selected, start=start + 1):
            numbered.append(f"{i}\t{line}")
        return "\n".join(numbered)
    except Exception as e:
        return f"Error: {e}"


def exec_edit(args: dict, config: AgentConfig) -> str:
    """Replace exact string in a file."""
    try:
        path = _resolve_in_jail(config, args["file_path"])
        if not path.exists():
            return f"Error: File not found: {path}"

        content = path.read_text()
        old = args["old_string"]
        new = args["new_string"]
        replace_all = args.get("replace_all", False)

        count = content.count(old)
        if count == 0:
            return f"Error: old_string not found in {path}"
        if count > 1 and not replace_all:
            return f"Error: old_string not unique in {path} (found {count} occurrences). Use replace_all=true to replace all."

        if replace_all:
            new_content = content.replace(old, new)
        else:
            new_content = content.replace(old, new, 1)

        path.write_text(new_content)
        return f"Replaced {count if replace_all else 1} occurrence(s) in {path}"
    except Exception as e:
        return f"Error: {e}"


def exec_write(args: dict, config: AgentConfig) -> str:
    """Write content to a file, creating parent dirs if needed."""
    try:
        path = _resolve_in_jail(config, args["file_path"])
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(args["content"])
        return f"Wrote {len(args['content'])} bytes to {path}"
    except Exception as e:
        return f"Error: {e}"
